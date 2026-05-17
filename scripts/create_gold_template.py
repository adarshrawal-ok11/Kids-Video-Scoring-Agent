"""One-time script to create data/gold_set_human.xlsx template."""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

os.makedirs("data", exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = "Gold Set"

columns = [
    ("id",                        "Unique ID e.g. gold_01",              14),
    ("url",                       "YouTube URL",                          45),
    ("auto_title",                "Auto-filled by system — leave blank",  35),
    ("auto_channel",              "Auto-filled by system — leave blank",  25),
    ("auto_duration_sec",         "Auto-filled by system — leave blank",  20),
    ("violence_aggression",       "0-100  (100 = no violence)",           22),
    ("scary_fear_inducing",       "0-100  (100 = not scary)",             22),
    ("inappropriate_language",    "0-100  (100 = clean language)",        24),
    ("bullying_exclusion",        "0-100  (100 = no bullying)",           22),
    ("stereotyping",              "0-100  (100 = no stereotypes)",        22),
    ("fantasy_level",             "0-100  (100 = realistic)",             18),
    ("negative_behavior_modeling","0-100  (100 = good behavior)",         26),
    ("visual_pacing",             "0-100  (100 = slow / calm)",           18),
    ("audio_energy",              "0-100  (100 = quiet / calm)",          18),
    ("color_intensity",           "0-100  (100 = muted colors)",          20),
    ("concepts_taught",           "0-100  (100 = very educational)",      22),
    ("participation",             "0-100  (100 = interactive)",           18),
    ("social_modeling",           "0-100  (100 = great modeling)",        20),
    ("repetition",                "0-100  (100 = good repetition)",       20),
    ("language_quality",          "0-100  (100 = rich language)",         20),
    ("content_consistency",       "0-100  (100 = consistent kids content)",26),
    ("channel_reputation",        "0-100  (100 = trusted channel)",       22),
    ("verdict",                   "APPROVED / REVIEW / REJECTED",         20),
    ("overall_estimate",          "Your rough overall score 0-100",       22),
    ("notes",                     "Anything notable about this video",    40),
    ("confidence",                "HIGH / MED / LOW",                     14),
    ("rated_by",                  "Your name",                            18),
    ("rated_on",                  "Date rated YYYY-MM-DD",                20),
    ("is_anchor",                 "TRUE for 5 anchor videos only",        20),
]

DARK_BLUE  = PatternFill("solid", fgColor="1F4E79")
GREY       = PatternFill("solid", fgColor="7F7F7F")
BLUE       = PatternFill("solid", fgColor="2E75B6")
DARK_GREEN = PatternFill("solid", fgColor="375623")
BROWN      = PatternFill("solid", fgColor="843C0C")
LIGHT_BLUE = PatternFill("solid", fgColor="D9E1F2")
ROW_FILL   = PatternFill("solid", fgColor="EBF3FB")

META_COLS   = {"id", "url", "verdict", "overall_estimate", "notes", "confidence", "rated_by", "rated_on"}
AUTO_COLS   = {"auto_title", "auto_channel", "auto_duration_sec"}

def header_fill(name: str) -> PatternFill:
    if name in AUTO_COLS:
        return GREY
    if name == "is_anchor":
        return BROWN
    if name in META_COLS:
        return DARK_GREEN
    return BLUE


for col_idx, (name, desc, width) in enumerate(columns, start=1):
    letter = get_column_letter(col_idx)
    ws.column_dimensions[letter].width = width

    h = ws.cell(row=1, column=col_idx, value=name)
    h.font = Font(bold=True, color="FFFFFF", size=10)
    h.fill = header_fill(name)
    h.alignment = Alignment(horizontal="center")

    d = ws.cell(row=2, column=col_idx, value=desc)
    d.font = Font(italic=True, color="404040", size=9)
    d.fill = LIGHT_BLUE
    d.alignment = Alignment(horizontal="center", wrap_text=True)

ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 30

col_names = [c[0] for c in columns]

SAMPLES = [
    {
        "id": "gold_01",
        "url": "https://www.youtube.com/watch?v=REPLACE_WITH_REAL_URL",
        "violence_aggression": 100, "scary_fear_inducing": 100,
        "inappropriate_language": 100, "bullying_exclusion": 100,
        "stereotyping": 90, "fantasy_level": 70,
        "negative_behavior_modeling": 85, "visual_pacing": 90,
        "audio_energy": 85, "color_intensity": 75,
        "concepts_taught": 90, "participation": 85,
        "social_modeling": 95, "repetition": 80,
        "language_quality": 85, "content_consistency": 95,
        "channel_reputation": 90,
        "verdict": "APPROVED", "overall_estimate": 88,
        "notes": "Sample APPROVED row — calm educational video, great for ages 3-5",
        "confidence": "HIGH", "rated_by": "Tara Team",
        "rated_on": "2026-05-13", "is_anchor": True,
    },
    {
        "id": "gold_02",
        "url": "https://www.youtube.com/watch?v=REPLACE_WITH_REAL_URL",
        "violence_aggression": 85, "scary_fear_inducing": 80,
        "inappropriate_language": 90, "bullying_exclusion": 75,
        "stereotyping": 70, "fantasy_level": 50,
        "negative_behavior_modeling": 60, "visual_pacing": 40,
        "audio_energy": 45, "color_intensity": 35,
        "concepts_taught": 60, "participation": 55,
        "social_modeling": 65, "repetition": 70,
        "language_quality": 60, "content_consistency": 80,
        "channel_reputation": 75,
        "verdict": "REVIEW", "overall_estimate": 62,
        "notes": "Sample REVIEW row — fast-paced bright colors but safe content",
        "confidence": "MED", "rated_by": "Tara Team",
        "rated_on": "2026-05-13", "is_anchor": False,
    },
    {
        "id": "gold_03",
        "url": "https://www.youtube.com/watch?v=REPLACE_WITH_REAL_URL",
        "violence_aggression": 30, "scary_fear_inducing": 25,
        "inappropriate_language": 40, "bullying_exclusion": 35,
        "stereotyping": 50, "fantasy_level": 20,
        "negative_behavior_modeling": 30, "visual_pacing": 20,
        "audio_energy": 25, "color_intensity": 20,
        "concepts_taught": 20, "participation": 15,
        "social_modeling": 20, "repetition": 30,
        "language_quality": 25, "content_consistency": 40,
        "channel_reputation": 30,
        "verdict": "REJECTED", "overall_estimate": 28,
        "notes": "Sample REJECTED row — aggressive content, not appropriate",
        "confidence": "HIGH", "rated_by": "Tara Team",
        "rated_on": "2026-05-13", "is_anchor": False,
    },
]

for row_idx, sample in enumerate(SAMPLES, start=3):
    for col_idx, name in enumerate(col_names, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=sample.get(name, ""))
        cell.alignment = Alignment(horizontal="center")
        cell.fill = ROW_FILL

# Dropdowns
verdict_col = get_column_letter(col_names.index("verdict") + 1)
dv_verdict = DataValidation(type="list", formula1='"APPROVED,REVIEW,REJECTED"', allow_blank=True)
dv_verdict.sqref = f"{verdict_col}3:{verdict_col}200"
ws.add_data_validation(dv_verdict)

conf_col = get_column_letter(col_names.index("confidence") + 1)
dv_conf = DataValidation(type="list", formula1='"HIGH,MED,LOW"', allow_blank=True)
dv_conf.sqref = f"{conf_col}3:{conf_col}200"
ws.add_data_validation(dv_conf)

anchor_col = get_column_letter(col_names.index("is_anchor") + 1)
dv_anchor = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
dv_anchor.sqref = f"{anchor_col}3:{anchor_col}200"
ws.add_data_validation(dv_anchor)

ws.freeze_panes = "B3"

out = Path("data/gold_set_human.xlsx")
wb.save(out)
print(f"Created {out}")
